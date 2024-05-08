#This app ranomly chooses one person to answer questions
#Out of cell A2 to A19 from excel file named hello that has all
#the student names from the class, we want ONE name to be randomly chosen.
import tkinter as tk
from tkinter.ttk import *
from openpyxl import load_workbook
import random

root = tk.Tk()
root.title('Lottery')
root.geometry("300x300")
root.resizable(0, 0)
root.grid_columnconfigure(1, weight=1)

def randnum():
    try:
        wb = load_workbook('hello.xlsx')
        ws = wb.active
        cell_range = ws['A2':'A19']
        students = [cell[0].value for cell in cell_range]
        winner = random.choice(students)
        result_label.config(text=winner)
    except FileNotFoundError:
        result_label.config(text="File not found")
    except Exception as e:
        result_label.config(text="Error: " + str(e))

b1 = tk.Button(text="Select Student Randomly", font=("Arial", 15), bg="#A3E4D7",
               command=randnum)
b1.grid(row=2, column=1)

result_label = tk.Label(bg="#F39C12", font=("Arial", 10), text="Result will display here")
result_label.grid(row=4, column=1)

tk.mainloop()
